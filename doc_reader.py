import os
import re
from docx import Document
import openpyxl

def get_competence_doc(file_name):
    document = Document(file_name)
    competence_dictionary = ""
    start_flag = False

    code = re.search("\d\d.\d\d.\d\d",file_name)
    if (code):
        code = code.group(0)

        # убрать в настройки
        str_start = 'V. ТРЕБОВАНИЯ К РЕЗУЛЬТАТАМ ОСВОЕНИЯ ПРОГРАММ'
        str_finish = 'VI. ТРЕБОВАНИЯ К СТРУКТУРЕ ПРОГРАММ'
        standart_errors = ['специалистов среднего звена', 'квалифицированных рабочих, служащих']

        if (code == '16.03.03'):
            a = 1
            a+= 1

        competence_list = []
        for para in document.paragraphs:
            if ((str_start.lower() in para.text.lower()) or (str_finish.lower() in para.text.lower())) :
                start_flag = not start_flag
            elif (start_flag and para.text != "" and not (para.text.lower() in standart_errors)) :
                competence_list.append(para.text)

        if (competence_list):
            competence_dictionary={code:competence_list}
    return competence_dictionary


def parse_folder(dir):
#dir = "/home/grigory/Desktop/АРМИЯ/ФГОСЫ"
    all_fgos_list = []

    subs = os.walk(dir)
    for d, dirs, files in subs:
        for f in files:
            file_name = d+'/'+f
            if (re.match(".*.docx",file_name)):
                competences = get_competence_doc(file_name)
                if (competences):
                    all_fgos_list.append(competences)
    return all_fgos_list


def docx_to_excel(all_fgos_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ФГОС"
    index = 1
    ws.cell(row=index, column=1).value = "ФГОС"
    ws.cell(row=index, column=2).value = "Компетенции"
    ws.cell(row=index, column=3).value = "Примечание"
    index += 1
    for fgos in all_fgos_list:
        for key in fgos.keys():
            ws.cell(row=index, column=1).value = key
            tmp_str = ""
            for value in fgos[key]:
                tmp_str += value + "\n"
            tmp_str = tmp_str.strip('\n')
        ws.cell(row=index, column=2).value = tmp_str
        ws.cell(row=index, column=3).value = ""
        index +=1
    wb.save("FGOS.xlsx")



dir = "/home/grigory/Desktop/АРМИЯ/ФГОСЫ"
docx_to_excel(parse_folder(dir))